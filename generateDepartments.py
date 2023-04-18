# -*- coding: utf-8 -*-
"""
Created on Wed Apr 12 12:36:42 2023

@author: Timo
"""

import openpyxl
from openpyxl.utils import column_index_from_string
from datetime import datetime, timedelta
import random

def generate_department_data(num_departments, num_weeks):
    # Generate the department names
    department_names = ["Department " + str(i+1) for i in range(num_departments)]

    # Generate the week numbers and dates
    start_date = datetime(2022, 1, 1)
    week_numbers = ["KW" + str(i).zfill(2) for i in range(1, num_weeks+1)]
    week_dates = [start_date + timedelta(weeks=i) for i in range(num_weeks)]

    # Generate the data for each cell
    data = []
    for i in range(num_weeks):
        week_data = []
        for j in range(num_departments):
            # Generate a random number between 6000 and 0
            value = random.randint(0, 6000)
            week_data.append(value)
        data.append(week_data)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Write the department names to the first row
    for i, name in enumerate(department_names):
        col_letter = openpyxl.utils.get_column_letter(i+2)
        sheet[col_letter + "1"] = name

    # Write the week numbers and dates to the first column
    for i, week_number in enumerate(week_numbers):
        sheet["A" + str(i+2)] = week_number
        sheet["B" + str(i+2)] = week_dates[i].strftime("%Y-%m-%d")

    # Write the data to the remaining cells
    for i, week_data in enumerate(data):
        for j, value in enumerate(week_data):
            col_letter = openpyxl.utils.get_column_letter(j+2)
            sheet[col_letter + str(i+2)] = value

    # Save the workbook
    wb.save("department_data.xlsx")

# Example usage: generate an Excel file with 10 departments and 200 weeks
generate_department_data(10, 200)